import openpyxl
from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\ACER\Downloads\TACH\docs\File Quan ly tong 2026.xlsx", data_only=True)

print("=== DANH SÁCH SHEET ===")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  [{ws.sheet_state}] {name}  (dim: {ws.dimensions})")

print()

target_sheets = [
    "DSLop", "DSHV", "DSTest", "ChiTietLopHoc",
    "TheoDoiHP", "XuatNhapSach", "Thu-Chi",
    "NhanSu", "Report_Cong_Luong", "Report_HP", "Report_HS", "MucLuc"
]

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"\n=== {sheet_name} [KHÔNG TỒN TẠI] ===")
        continue
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"=== SHEET: {sheet_name} (state={ws.sheet_state}, dim={ws.dimensions}) ===")
    print(f"{'='*60}")

    # Print first 60 rows, max 40 cols
    max_row = min(ws.max_row or 0, 60)
    max_col = min(ws.max_column or 0, 40)

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            else:
                s = str(v).replace('\n', ' ').strip()
                if len(s) > 30:
                    s = s[:30] + "…"
                cells.append(s)
        print(" | ".join(cells))
