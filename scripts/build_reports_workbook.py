import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
CARD_FILL = PatternFill("solid", fgColor="EAF4EA")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def vnd_format():
    return '#,##0"đ"'


def style_range(ws, cell_range, fill=None, font=None, align=None, border=True):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if border:
                cell.border = THIN_BORDER


def autofit(ws, widths):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_report_hs(ws, data):
    ws.title = "Report_HS"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    ws["A1"] = "REPORT_HS - ERP"
    ws["A2"] = "Đối soát học viên theo dữ liệu thật từ ERP"
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    style_range(ws, "A1:F1", fill=HEADER_FILL, font=Font(color="FFFFFF", bold=True, size=14), align=Alignment(horizontal="center"))
    style_range(ws, "A2:F2", fill=SECTION_FILL, font=BOLD_FONT, align=Alignment(horizontal="center"))

    summary_headers = ["Chỉ số", "Giá trị"]
    summary_rows = [
        ["Tổng số HS", data["totalStudents"]],
        ["Số HS đang học", data["activeStudents"]],
        ["Số HS nghỉ", data["leftStudents"]],
        ["Số HS nhập học", data["newEnrollments"]],
    ]
    ws["A4"] = "Tóm tắt"
    style_range(ws, "A4:B4", fill=SECTION_FILL, font=BOLD_FONT)
    for col, value in enumerate(summary_headers, start=1):
        ws.cell(5, col).value = value
    style_range(ws, "A5:B5", fill=HEADER_FILL, font=WHITE_FONT, align=Alignment(horizontal="center"))
    for row_idx, row in enumerate(summary_rows, start=6):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx).value = value
            ws.cell(row_idx, col_idx).border = THIN_BORDER
    style_range(ws, "A6:B9", fill=CARD_FILL)

    table_row = 12
    ws[f"A{table_row}"] = "Theo lớp"
    style_range(ws, f"A{table_row}:E{table_row}", fill=SECTION_FILL, font=BOLD_FONT)
    headers = ["Mã lớp", "Tên lớp", "Đang học", "Đã nghỉ", "Tổng"]
    for col_idx, value in enumerate(headers, start=1):
        ws.cell(table_row + 1, col_idx).value = value
    style_range(ws, f"A{table_row + 1}:E{table_row + 1}", fill=HEADER_FILL, font=WHITE_FONT, align=Alignment(horizontal="center"))

    current_row = table_row + 2
    for item in data["classes"]:
        ws.cell(current_row, 1).value = item["classCode"]
        ws.cell(current_row, 2).value = item["className"]
        ws.cell(current_row, 3).value = item["activeCount"]
        ws.cell(current_row, 4).value = item["leftCount"]
        ws.cell(current_row, 5).value = item["totalCount"]
        style_range(ws, f"A{current_row}:E{current_row}")
        current_row += 1

    autofit(ws, {1: 16, 2: 34, 3: 14, 4: 14, 5: 12})


def build_report_hp(ws, data):
    ws.title = "Report_HP"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws["A1"] = "REPORT_HP - ERP"
    ws["A2"] = "Đối soát học phí theo dữ liệu thật từ ERP"
    ws["A3"] = f"Kỳ học phí: {data['periodName'] or 'Chưa có kỳ'}"
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws.merge_cells("A3:H3")
    style_range(ws, "A1:H1", fill=HEADER_FILL, font=Font(color="FFFFFF", bold=True, size=14), align=Alignment(horizontal="center"))
    style_range(ws, "A2:H2", fill=SECTION_FILL, font=BOLD_FONT, align=Alignment(horizontal="center"))
    style_range(ws, "A3:H3", fill=CARD_FILL, font=BOLD_FONT, align=Alignment(horizontal="center"))

    ws["A5"] = "Tổng hợp"
    style_range(ws, "A5:H5", fill=SECTION_FILL, font=BOLD_FONT)
    headers = [
        "Số buổi",
        "Giáo trình",
        "HP tồn tháng trước",
        "Tổng học phí",
        "Phải thu",
        "Đã thu",
        "Còn lại",
    ]
    values = [
        data["totals"]["sessionCount"],
        data["totals"]["materialsAmount"],
        data["totals"]["openingBalance"],
        data["totals"]["tuitionAmount"],
        data["totals"]["billedAmount"],
        data["totals"]["collectedAmount"],
        data["totals"]["remainingAmount"],
    ]

    for idx, value in enumerate(headers, start=1):
        ws.cell(6, idx).value = value
    style_range(ws, "A6:G6", fill=HEADER_FILL, font=WHITE_FONT, align=Alignment(horizontal="center"))
    for idx, value in enumerate(values, start=1):
        ws.cell(7, idx).value = value
        ws.cell(7, idx).border = THIN_BORDER
        if idx >= 2:
            ws.cell(7, idx).number_format = vnd_format()
    style_range(ws, "A7:G7", fill=CARD_FILL)

    table_row = 10
    ws[f"A{table_row}"] = "Theo lớp"
    style_range(ws, f"A{table_row}:I{table_row}", fill=SECTION_FILL, font=BOLD_FONT)
    class_headers = [
        "Mã lớp",
        "Tên lớp",
        "Số HV",
        "Số buổi",
        "Giáo trình",
        "HP tồn tháng trước",
        "Tổng học phí",
        "Phải thu",
        "Đã thu",
        "Còn lại",
    ]
    for col_idx, value in enumerate(class_headers, start=1):
        ws.cell(table_row + 1, col_idx).value = value
    style_range(ws, f"A{table_row + 1}:J{table_row + 1}", fill=HEADER_FILL, font=WHITE_FONT, align=Alignment(horizontal="center"))

    current_row = table_row + 2
    for item in data["classes"]:
        row_values = [
            item["classCode"],
            item["className"],
            item["studentCount"],
            item["sessionCount"],
            item["materialsAmount"],
            item["openingBalance"],
            item["tuitionAmount"],
            item["billedAmount"],
            item["collectedAmount"],
            item["remainingAmount"],
        ]
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(current_row, col_idx).value = value
            ws.cell(current_row, col_idx).border = THIN_BORDER
            if col_idx >= 5:
                ws.cell(current_row, col_idx).number_format = vnd_format()
        current_row += 1

    autofit(ws, {1: 14, 2: 34, 3: 10, 4: 10, 5: 14, 6: 18, 7: 16, 8: 16, 9: 16, 10: 14})


def main():
    payload_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    payload = json.loads(payload_path.read_text(encoding="utf-8"))

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    build_report_hs(workbook.create_sheet(), payload["reportHs"])
    build_report_hp(workbook.create_sheet(), payload["reportHp"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(json.dumps({"ok": True, "output": str(output_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
