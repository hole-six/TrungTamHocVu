from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

ALLOWED_TABLE_KEYS = {
    ("ChiTietLopHoc", "T_ChiTietLop"),
    ("DSTest", "T_DSTest"),
    ("DSHV", "T_HV"),
    ("TheoDoiHP", "T_HP"),
    ("XuatNhapSach", "T_SachTon"),
    ("XuatNhapSach", "T_SachNhap"),
    ("XuatNhapSach", "T_SachXuat"),
    ("Thu-Chi", "T_Thu"),
    ("Thu-Chi", "T_Chi"),
    ("Thu-Chi", "T_PhanLoai"),
    ("DSLop", "T_DSLop"),
    ("NhanSu", "T_NS"),
    ("MucLuc", "Table1"),
    ("MucLuc", "Table2"),
    ("MucLuc", "Table3"),
}

BUSINESS_KEY_FIELDS = {
    "DSTest.T_DSTest": ["MaSo", "HoTenHV", "Sdt"],
    "DSHV.T_HV": ["MaSo", "TenHV", "MaHV"],
    "DSLop.T_DSLop": ["MaLop", "Ten lop", "TenLop"],
    "NhanSu.T_NS": ["Mã NV", "Họ và tên", "Tên ngắn"],
    "TheoDoiHP.T_HP": ["MaSo", "Ten HV", "HP Tháng"],
    "XuatNhapSach.T_SachTon": ["TenSach"],
    "XuatNhapSach.T_SachNhap": ["TenSach", "Ngày tháng"],
    "XuatNhapSach.T_SachXuat": ["TenSach", "TenHV&MaHV", "NgayThang"],
    "Thu-Chi.T_Thu": ["Ngày tháng", "Loại thu", "Số tiền"],
    "Thu-Chi.T_Chi": ["Ngày tháng", "Loại chi", "Số tiền"],
    "Thu-Chi.T_PhanLoai": ["LoaiHinh", "TenThuChi"],
    "MucLuc.Table1": ["KhungTG", "ThoiGian"],
    "MucLuc.Table2": ["MaLop", "TenLop"],
    "MucLuc.Table3": ["MaThu", "TenThu"],
    "ChiTietLopHoc.T_ChiTietLop": ["Ngay thang", "MaLop", "TenLop", "Giáo viên"],
}

PLACEHOLDER_TEXTS = {
    "",
    "-",
    "(blank)",
    "#n/a",
    "#ref!",
    "#value!",
    "0",
    "0.0",
    "00:00:00",
    "1/1900",
    "1900-1",
    "3-",
    ".3-",
    "chưa có info",
    "đã tt đủ",
    "đóng đủ",
    "liên hệ ngay",
    "chưa đi học",
    "đã nhập dshv",
}


@dataclass
class FieldSpec:
    name: str
    inferred_type: str
    classification: str
    required: str
    validation: str
    source_status: str


@dataclass
class TableSpec:
    sheet: str
    table: str
    cell_range: str
    fields: list[FieldSpec]


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else round(value, 4)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
        return int(number) if number.is_integer() else round(number, 4)
    except ValueError:
        return None


def normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None

    patterns = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
    ]
    for pattern in patterns:
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def normalize_value(value: Any, inferred_type: str) -> Any:
    lower_type = inferred_type.lower()
    if "date" in lower_type:
        return normalize_date(value)
    if "number" in lower_type:
        number = normalize_number(value)
        return number if number is not None else normalize_text(value)
    return normalize_text(value)


def is_placeholder_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, (int, float)) and value == 0:
        return True
    normalized = normalize_text(value)
    if normalized is None:
        return True
    return normalized.lower() in PLACEHOLDER_TEXTS


def table_to_filename(table_name: str) -> str:
    return re.sub(r"[^0-9A-Za-z]+", "_", table_name).strip("_") + ".csv"


def slugify(text: str | None) -> str:
    if not text:
        return ""
    normalized = re.sub(r"[^0-9a-z]+", "-", text.lower()).strip("-")
    return normalized


def month_key(value: Any) -> str | None:
    normalized = normalize_date(value)
    if normalized and re.match(r"^\d{4}-\d{2}-\d{2}$", normalized):
        return normalized[:7]

    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("\\", "/").replace(".", "/").strip()
    patterns = [
        r"^(?P<year>\d{4})[-/](?P<month>\d{1,2})$",
        r"^(?P<month>\d{1,2})[-/](?P<year>\d{4})$",
        r"^(?P<month>\d{1,2})[-/](?P<year>\d{2})$",
    ]
    for pattern in patterns:
        match = re.match(pattern, text)
        if not match:
            continue
        month = int(match.group("month"))
        year = int(match.group("year"))
        year = year + 2000 if year < 100 else year
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}"
    return text


def parse_time_range(value: Any) -> tuple[str | None, str | None]:
    text = normalize_text(value)
    if not text:
        return None, None
    match = re.search(r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})", text)
    if match:
        return match.group(1), match.group(2)
    return None, None


def normalize_status(text: Any, mapping: dict[str, str], fallback: str) -> str:
    value = normalize_text(text)
    if not value:
        return fallback
    key = value.lower()
    for candidate, result in mapping.items():
        if candidate in key:
            return result
    return fallback


def read_dictionary(path: Path) -> list[TableSpec]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            key = (row["sheet"], row["table"])
            if key not in ALLOWED_TABLE_KEYS:
                continue
            grouped.setdefault(
                key,
                {
                    "sheet": row["sheet"],
                    "table": row["table"],
                    "cell_range": row["range"],
                    "fields": [],
                },
            )
            grouped[key]["fields"].append(
                (
                    int(row["column_index"]),
                    FieldSpec(
                        name=row["field_name"],
                        inferred_type=row["inferred_type"],
                        classification=row["classification"],
                        required=row["required"],
                        validation=row["validation"],
                        source_status=row["source_status"],
                    ),
                )
            )

    specs: list[TableSpec] = []
    for group in grouped.values():
        ordered_fields = [item[1] for item in sorted(group["fields"], key=lambda value: value[0])]
        specs.append(
            TableSpec(
                sheet=group["sheet"],
                table=group["table"],
                cell_range=group["cell_range"],
                fields=ordered_fields,
            )
        )
    return sorted(specs, key=lambda item: (item.sheet, item.table))


def extract_tables(workbook_path: Path, specs: list[TableSpec]) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    extracted: dict[str, list[dict[str, Any]]] = {}

    for spec in specs:
        worksheet = workbook[spec.sheet]
        min_col, min_row, max_col, max_row = range_boundaries(spec.cell_range)
        rows: list[dict[str, Any]] = []
        for row_number, raw_row in enumerate(
            worksheet.iter_rows(
            min_row=min_row + 1,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
            ),
            start=min_row + 1,
        ):
            record: dict[str, Any] = {}
            has_data = False
            for offset, field in enumerate(spec.fields):
                if offset >= len(raw_row):
                    break
                value = raw_row[offset]
                normalized = normalize_value(value, field.inferred_type)
                if normalized not in (None, ""):
                    has_data = True
                record[field.name] = normalized
            if has_data:
                record["__sourceRow"] = row_number
                rows.append(record)
        extracted[f"{spec.sheet}.{spec.table}"] = rows
    return extracted


def build_field_type_lookup(specs: list[TableSpec]) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for spec in specs:
        table_name = f"{spec.sheet}.{spec.table}"
        lookup[table_name] = {field.name: field.inferred_type for field in spec.fields}
    return lookup


def load_csv_overrides(
    remediation_dir: Path,
    table_types: dict[str, dict[str, str]],
) -> tuple[dict[str, dict[int, dict[str, Any]]], dict[str, Any]]:
    overrides: dict[str, dict[int, dict[str, Any]]] = {}
    summary = {
        "filesRead": 0,
        "rowsApplied": 0,
        "tablesTouched": 0,
    }

    if not remediation_dir.exists():
        return overrides, summary

    filename_to_table = {
        table_to_filename(table_name): table_name for table_name in table_types.keys()
    }

    for csv_path in remediation_dir.glob("*.csv"):
        if csv_path.name.startswith("_"):
            continue
        table_name = filename_to_table.get(csv_path.name)
        if not table_name:
            continue
        file_rows_applied = 0
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                apply_flag = normalize_text(row.get("applyOverride"))
                if not apply_flag or apply_flag.lower() not in {"1", "y", "yes", "true", "x"}:
                    continue
                source_row_text = normalize_text(row.get("sourceRow"))
                if not source_row_text or not source_row_text.isdigit():
                    continue
                source_row = int(source_row_text)
                patch: dict[str, Any] = {}
                for field_name, value in row.items():
                    if field_name in {"sourceRow", "applyOverride", "notes"}:
                        continue
                    normalized = normalize_text(value)
                    if normalized is None:
                        continue
                    inferred_type = table_types.get(table_name, {}).get(field_name, "text")
                    patch[field_name] = normalize_value(normalized, inferred_type)
                if not patch:
                    continue
                overrides.setdefault(table_name, {})[source_row] = patch
                file_rows_applied += 1

        if file_rows_applied:
            summary["filesRead"] += 1
            summary["rowsApplied"] += file_rows_applied
            summary["tablesTouched"] += 1

    return overrides, summary


def apply_overrides(
    raw_tables: dict[str, list[dict[str, Any]]],
    overrides: dict[str, dict[int, dict[str, Any]]],
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    merged: dict[str, list[dict[str, Any]]] = {}
    applied_count = 0
    touched_tables = 0

    for table_name, rows in raw_tables.items():
        row_overrides = overrides.get(table_name, {})
        if not row_overrides:
            merged[table_name] = rows
            continue

        touched_tables += 1
        merged_rows: list[dict[str, Any]] = []
        for row in rows:
            source_row = row.get("__sourceRow")
            if isinstance(source_row, int) and source_row in row_overrides:
                merged_rows.append({**row, **row_overrides[source_row]})
                applied_count += 1
            else:
                merged_rows.append(row)
        merged[table_name] = merged_rows

    return merged, {
        "rowsApplied": applied_count,
        "tablesTouched": touched_tables,
    }


def build_diagnostics(raw_tables: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    table_reports: dict[str, Any] = {}
    overall = Counter()

    for table_name, rows in raw_tables.items():
        key_fields = BUSINESS_KEY_FIELDS.get(table_name, [])
        rows_with_any_key = 0
        rows_with_all_keys = 0
        placeholder_only_rows = 0
        meaningful_rows = 0
        samples: list[dict[str, Any]] = []
        missing_key_counter = Counter()

        for row in rows:
            key_values = {field: row.get(field) for field in key_fields}
            valid_key_fields = [field for field, value in key_values.items() if not is_placeholder_value(value)]
            if valid_key_fields:
                rows_with_any_key += 1
            if key_fields and len(valid_key_fields) == len(key_fields):
                rows_with_all_keys += 1
            if key_fields:
                for field in key_fields:
                    if is_placeholder_value(row.get(field)):
                        missing_key_counter[field] += 1

            non_meta_items = {
                key: value
                for key, value in row.items()
                if key != "__sourceRow" and not is_placeholder_value(value)
            }
            if non_meta_items:
                if any(
                    value.lower() not in PLACEHOLDER_TEXTS
                    for value in (normalize_text(item) or "" for item in non_meta_items.values())
                ):
                    meaningful_rows += 1
            else:
                placeholder_only_rows += 1

            if len(samples) < 5 and valid_key_fields:
                samples.append(
                    {
                        "sourceRow": row.get("__sourceRow"),
                        "keys": {field: row.get(field) for field in key_fields},
                    }
                )

        row_count = len(rows)
        quality_status = "EMPTY_KEYS"
        if rows_with_any_key == 0:
            quality_status = "PLACEHOLDER_ONLY"
        elif rows_with_all_keys > 0:
            quality_status = "READY_FOR_IMPORT"
        elif rows_with_any_key > 0:
            quality_status = "PARTIAL_KEYS"

        report = {
            "rowCount": row_count,
            "businessKeyFields": key_fields,
            "rowsWithAnyKey": rows_with_any_key,
            "rowsWithAllKeys": rows_with_all_keys,
            "placeholderOnlyRows": placeholder_only_rows,
            "meaningfulRows": meaningful_rows,
            "qualityStatus": quality_status,
            "missingKeyCounts": dict(missing_key_counter),
            "sampleRowsWithKeys": samples,
        }
        table_reports[table_name] = report
        overall["tables"] += 1
        if quality_status == "READY_FOR_IMPORT":
            overall["readyTables"] += 1
        elif quality_status == "PARTIAL_KEYS":
            overall["partialTables"] += 1
        else:
            overall["placeholderTables"] += 1

    return {"overall": dict(overall), "tables": table_reports}


def canonicalize(raw_tables: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    summary: dict[str, Any] = {
        "warnings": [],
        "linkStats": {},
    }

    raw_courses = raw_tables.get("MucLuc.Table2", [])
    raw_employees = raw_tables.get("NhanSu.T_NS", [])
    raw_leads = raw_tables.get("DSTest.T_DSTest", [])
    raw_students = raw_tables.get("DSHV.T_HV", [])
    raw_classes = raw_tables.get("DSLop.T_DSLop", [])
    raw_tuition = raw_tables.get("TheoDoiHP.T_HP", [])
    raw_books = raw_tables.get("XuatNhapSach.T_SachTon", [])
    raw_receipts = raw_tables.get("XuatNhapSach.T_SachNhap", [])
    raw_issues = raw_tables.get("XuatNhapSach.T_SachXuat", [])
    raw_cash_categories = raw_tables.get("Thu-Chi.T_PhanLoai", [])
    raw_cash_in = raw_tables.get("Thu-Chi.T_Thu", [])
    raw_cash_out = raw_tables.get("Thu-Chi.T_Chi", [])
    raw_weekdays = raw_tables.get("MucLuc.Table3", [])
    raw_time_slots = raw_tables.get("MucLuc.Table1", [])

    courses: list[dict[str, Any]] = []
    course_by_code: dict[str, dict[str, Any]] = {}
    for row in raw_courses:
        code = normalize_text(row.get("MaLop"))
        name = normalize_text(row.get("TenLop"))
        if not code or not name:
            continue
        item = {
            "courseCode": code,
            "name": name,
            "tuitionPerSession": normalize_number(row.get("HocPhi/Buoi")),
            "sessionsPerWeek": normalize_number(row.get("BuoiHoc/Tuan")),
            "businessKey": code,
        }
        course_by_code[code] = item
        courses.append(item)

    employees: list[dict[str, Any]] = []
    employee_by_short_name: dict[str, dict[str, Any]] = {}
    for row in raw_employees:
        employee_code = normalize_text(row.get("Mã NV"))
        short_name = normalize_text(row.get("Tên ngắn"))
        full_name = normalize_text(row.get("Họ và tên"))
        if not employee_code or not short_name or not full_name:
            continue
        raw_mode = normalize_text(row.get("Column4"))
        pay_mode = "SESSION" if raw_mode and "ca" in raw_mode.lower() else "HOURLY"
        item = {
            "employeeCode": employee_code,
            "fullName": full_name,
            "shortName": short_name,
            "dob": normalize_date(row.get("Ngày sinh")),
            "position": normalize_text(row.get("Vị trí")),
            "phone": normalize_text(row.get("SĐT")),
            "email": normalize_text(row.get("Mail")),
            "hometown": normalize_text(row.get("Quê quán")),
            "permanentAddress": normalize_text(row.get("Địa chỉ thường trú")),
            "idNumber": normalize_text(row.get("Số CMT")),
            "idIssueDate": normalize_date(row.get("Ngày cấp")),
            "idIssuePlace": normalize_text(row.get("Nơi cấp")),
            "contractSignDate": normalize_date(row.get("Ngày ký HĐ")),
            "contractExpiryDate": normalize_date(row.get("Hạn HĐ")),
            "resignDate": normalize_date(row.get("Ngày nghỉ")),
            "workStatus": "RESIGNED" if normalize_date(row.get("Ngày nghỉ")) else "ACTIVE",
            "teachingHourlyRate": normalize_number(row.get("Column3")),
            "payMode": pay_mode,
            "rawPayModeHint": raw_mode,
            "rawAssistantRateHint": normalize_text(row.get("Column4")),
        }
        employees.append(item)
        employee_by_short_name[short_name.lower()] = item

    guardians: list[dict[str, Any]] = []
    guardian_by_key: dict[str, dict[str, Any]] = {}
    leads: list[dict[str, Any]] = []
    lead_by_code: dict[str, dict[str, Any]] = {}
    lead_status_map = {
        "đã đi học": "ENROLLED",
        "đã nhập dshv": "ENROLLED",
        "chưa test": "CONTACTING",
        "test": "TESTED",
        "không": "LOST",
    }

    for index, row in enumerate(raw_leads, start=1):
        full_name = normalize_text(row.get("HoTenHV"))
        if not full_name:
            continue
        phone = normalize_text(row.get("Sdt"))
        guardian_name = normalize_text(row.get("HoTenPH"))
        guardian_key = f"{slugify(guardian_name)}|{phone or ''}" if guardian_name else ""
        guardian_ref = None
        if guardian_name:
            guardian_ref = guardian_by_key.get(guardian_key)
            if guardian_ref is None:
                guardian_ref = {
                    "businessKey": guardian_key or f"guardian-{len(guardians)+1}",
                    "fullName": guardian_name,
                    "phone": phone,
                    "address": normalize_text(row.get("DiaChiNha")),
                }
                guardian_by_key[guardian_key] = guardian_ref
                guardians.append(guardian_ref)

        lead_code = normalize_text(row.get("MaSo")) or f"DSTEST-{index:04d}"
        expected_start = normalize_date(row.get("Ngày dự kiến đi học"))
        actual_enroll = normalize_date(row.get("Ngày nhập học"))
        source_status = normalize_text(row.get("Tình trạng test"))
        proposed_status = "ENROLLED" if actual_enroll else normalize_status(source_status, lead_status_map, "NEW")

        interested_class_code = normalize_text(row.get("TenLop"))
        item = {
            "leadCode": lead_code,
            "fullName": full_name,
            "gender": normalize_text(row.get("GioiTinh")),
            "dob": normalize_date(row.get("DoB")),
            "phone": phone,
            "address": normalize_text(row.get("DiaChiNha")),
            "meetDate": normalize_date(row.get("NgayGap")),
            "interestedClassRef": interested_class_code,
            "status": proposed_status,
            "sourceStatus": source_status,
            "expectedStartDate": expected_start,
            "actualEnrollDate": actual_enroll,
            "notes": normalize_text(row.get("GhiChu")),
            "notes2": normalize_text(row.get("GhiChu2")),
            "guardianRef": guardian_ref["businessKey"] if guardian_ref else None,
            "sourceBusinessKey": normalize_text(row.get("TenHV&MaSo")),
        }
        leads.append(item)
        lead_by_code[lead_code] = item

    students: list[dict[str, Any]] = []
    student_by_code: dict[str, dict[str, Any]] = {}
    enrollments: list[dict[str, Any]] = []
    student_guardian_links: list[dict[str, Any]] = []
    for index, row in enumerate(raw_students, start=1):
        full_name = normalize_text(row.get("TenHV"))
        if not full_name:
            continue
        student_code = normalize_text(row.get("MaSo")) or f"DSHV-{index:04d}"
        student_display_id = normalize_text(row.get("MaHV"))
        class_code = normalize_text(row.get("MaLop"))
        class_name = normalize_text(row.get("TenLop"))
        leave_date = normalize_date(row.get("Ngay nghi"))
        item = {
            "studentCode": student_code,
            "studentDisplayId": student_display_id,
            "fullName": full_name,
            "phone": normalize_text(row.get("Sđt")),
            "enrollDate": normalize_date(row.get("Ngay nhap")),
            "leaveDate": leave_date,
            "leaveReason": normalize_text(row.get("Lí do nghỉ")),
            "evaluation": normalize_text(row.get("DanhGia")),
            "status": "LEFT" if leave_date else "ACTIVE",
            "leadRef": student_code if student_code in lead_by_code else None,
            "classRef": class_code or class_name,
            "sourceBusinessKey": normalize_text(row.get("TenHV&MaSo")) or normalize_text(row.get("TenHV&MaHV")),
        }
        students.append(item)
        student_by_code[student_code] = item

        if item["classRef"]:
            enrollments.append(
                {
                    "studentRef": student_code,
                    "classRef": item["classRef"],
                    "status": "WITHDRAWN" if leave_date else "ACTIVE",
                    "enrollDate": item["enrollDate"],
                    "endDate": leave_date,
                }
            )

        matched_guardian = None
        if item["leadRef"] and item["leadRef"] in lead_by_code:
            guardian_ref = lead_by_code[item["leadRef"]].get("guardianRef")
            matched_guardian = guardian_ref
        if matched_guardian:
            student_guardian_links.append(
                {
                    "studentRef": student_code,
                    "guardianRef": matched_guardian,
                    "isPrimary": True,
                }
            )

    classes: list[dict[str, Any]] = []
    class_by_ref: dict[str, dict[str, Any]] = {}
    for row in raw_classes:
        class_code = normalize_text(row.get("MaLop"))
        class_name = normalize_text(row.get("Ten lop")) or normalize_text(row.get("TenLop"))
        if not class_code or not class_name:
            continue
        course = course_by_code.get(class_code)
        item = {
            "classCode": class_code,
            "classGroup": normalize_text(row.get("Lop")),
            "className": class_name,
            "courseRef": course["courseCode"] if course else class_code,
            "totalSessions": normalize_number(row.get("SLBuoiHoc")),
            "startDate": normalize_date(row.get("NgayBD")),
            "expectedEndDate": normalize_date(row.get("NgayKTDuKien")),
            "sessionsPerWeek": normalize_number(row.get("BuoiHoc/Tuan")) or (course["sessionsPerWeek"] if course else None),
            "tuitionPerSession": normalize_number(row.get("HocPhi/Buoi")) or (course["tuitionPerSession"] if course else None),
            "status": "ACTIVE",
            "notes": normalize_text(row.get("GhiChu")) or normalize_text(row.get("ghi chu")),
            "reminderDayOfMonth": normalize_number(row.get("Ngay CĐ")),
            "reminderWeekday": normalize_text(row.get("Thu")),
            "reminderTask": normalize_text(row.get("Cong viec")),
            "oneOffTaskDate": normalize_date(row.get("Ngay phat sinh")),
            "oneOffTaskName": normalize_text(row.get("Cong viec PS")),
        }
        classes.append(item)
        class_by_ref[class_code] = item
        class_by_ref[class_name.lower()] = item

    billing_periods: list[dict[str, Any]] = []
    billing_period_seen: set[str] = set()
    charges: list[dict[str, Any]] = []
    payments: list[dict[str, Any]] = []
    unresolved_charge_links = 0
    for index, row in enumerate(raw_tuition, start=1):
        period_name = month_key(row.get("HP Tháng"))
        student_ref = normalize_text(row.get("MaSo"))
        class_ref = normalize_text(row.get("MaLop")) or normalize_text(row.get("TenLop"))
        if period_name and period_name not in billing_period_seen:
            billing_period_seen.add(period_name)
            billing_periods.append({"periodName": period_name})

        if student_ref and class_ref and period_name:
            charge_key = f"{student_ref}|{class_ref}|{period_name}"
            charges.append(
                {
                    "businessKey": charge_key,
                    "studentRef": student_ref,
                    "classRef": class_ref,
                    "billingPeriodRef": period_name,
                    "sessionCount": normalize_number(row.get("So buoi")) or 0,
                    "absentCount": normalize_number(row.get("Buoi nghi")) or 0,
                    "deductedCount": normalize_number(row.get("Buoi tru")) or 0,
                    "unitPrice": normalize_number(row.get("ĐG")) or 0,
                    "tuitionAmount": normalize_number(row.get("HP thang hien tai")) or 0,
                    "materialsAmount": normalize_number(row.get("TienGiaoTrinh")) or 0,
                    "openingBalance": normalize_number(row.get("HP dau ky")) or 0,
                    "totalAmount": normalize_number(row.get("TongHP")) or 0,
                    "closingBalance": normalize_number(row.get("Con lai")),
                    "paymentStatus": normalize_text(row.get("Tình trạng đóng học phí")),
                    "notes": normalize_text(row.get("GhiChu")),
                }
            )
        else:
            unresolved_charge_links += 1

        payment_amount = normalize_number(row.get("TienNop"))
        if student_ref and payment_amount not in (None, 0):
            payments.append(
                {
                    "paymentNo": f"HP-{index:05d}",
                    "studentRef": student_ref,
                    "paidDate": normalize_date(row.get("NgayNopTien")) or f"{period_name}-01" if period_name else None,
                    "amount": payment_amount,
                    "method": normalize_text(row.get("HinhThucTT")),
                    "receivedByName": normalize_text(row.get("Người nhận")),
                    "billingPeriodRef": period_name,
                    "notes": normalize_text(row.get("GhiChu")),
                }
            )

    books: list[dict[str, Any]] = []
    book_by_name: dict[str, dict[str, Any]] = {}
    for index, row in enumerate(raw_books, start=1):
        name = normalize_text(row.get("TenSach"))
        if not name:
            continue
        item = {
            "bookCode": normalize_text(row.get("Column1")) or f"BOOK-{index:04d}",
            "classRef": normalize_text(row.get("MaLop")),
            "name": name,
            "unitPrice": normalize_number(row.get("DonGia")) or 0,
            "quantityOnHand": normalize_number(row.get("Số Lượng")) or 0,
            "notes": normalize_text(row.get("GhiChu")),
        }
        books.append(item)
        book_by_name[name.lower()] = item

    stock_receipts: list[dict[str, Any]] = []
    for index, row in enumerate(raw_receipts, start=1):
        book_name = normalize_text(row.get("TenSach"))
        if not book_name:
            continue
        quantity = normalize_number(row.get("SL nhập")) or 0
        total_amount = normalize_number(row.get("Tổng tiền"))
        unit_price = normalize_number(row.get("Đơn giá")) or (
            int(total_amount / quantity) if total_amount and quantity else 0
        )
        stock_receipts.append(
            {
                "receiptNo": f"RECEIPT-{index:05d}",
                "bookRef": book_name,
                "classRef": normalize_text(row.get("Malop")),
                "quantity": quantity,
                "unitPrice": unit_price,
                "totalAmount": total_amount or (quantity * unit_price),
                "txnDate": normalize_date(row.get("Ngày tháng")),
                "receivedByName": normalize_text(row.get("Người nhập")),
                "handedByName": normalize_text(row.get("Người giao")),
                "usageStatus": normalize_text(row.get("Tình trạng sử dụng")),
                "notes": normalize_text(row.get("Ghi chú")),
                "billingPeriodRef": month_key(row.get("Tháng nhập")),
            }
        )

    book_issues: list[dict[str, Any]] = []
    unresolved_book_issue_links = 0
    student_by_display_key = {
        normalize_text(item.get("sourceBusinessKey") or item["studentCode"]): item
        for item in students
        if normalize_text(item.get("sourceBusinessKey") or item["studentCode"])
    }
    for index, row in enumerate(raw_issues, start=1):
        student_display_key = normalize_text(row.get("TenHV&MaHV"))
        student_ref = None
        if student_display_key:
            for candidate in students:
                candidate_key = normalize_text(candidate.get("sourceBusinessKey"))
                if candidate_key and candidate_key == student_display_key:
                    student_ref = candidate["studentCode"]
                    break
            if not student_ref:
                match = next((item for item in students if item.get("studentDisplayId") and item["studentDisplayId"] in student_display_key), None)
                if match:
                    student_ref = match["studentCode"]
        if not student_ref:
            unresolved_book_issue_links += 1
        book_issues.append(
            {
                "issueNo": f"ISSUE-{index:05d}",
                "bookRef": normalize_text(row.get("TenSach")),
                "classRef": normalize_text(row.get("MaLop")) or normalize_text(row.get("TenLop")),
                "studentRef": student_ref,
                "studentDisplayKey": student_display_key,
                "quantity": normalize_number(row.get("SL")) or 0,
                "unitPrice": normalize_number(row.get("DonGia")) or 0,
                "amount": normalize_number(row.get("TienGiaoTrinh")) or 0,
                "issueDate": normalize_date(row.get("NgayThang")),
                "billingPeriodRef": month_key(row.get("Tháng xuất")),
                "notes": normalize_text(row.get("GhiChu")),
            }
        )

    transaction_categories: list[dict[str, Any]] = []
    for row in raw_cash_categories:
        txn_type = normalize_text(row.get("LoaiHinh"))
        name = normalize_text(row.get("TenThuChi"))
        if not txn_type or not name:
            continue
        transaction_categories.append(
            {
                "type": txn_type.upper(),
                "name": name,
                "detail": normalize_text(row.get("ChiTietLoai")),
                "notes": normalize_text(row.get("GhiChu")),
                "handledByHint": normalize_text(row.get("nguoi chi")) or normalize_text(row.get("nguoi Thu")),
            }
        )

    cash_transactions: list[dict[str, Any]] = []
    for txn_type, rows in (("THU", raw_cash_in), ("CHI", raw_cash_out)):
        for index, row in enumerate(rows, start=1):
            cash_transactions.append(
                {
                    "transactionNo": f"{txn_type}-{index:05d}",
                    "type": txn_type,
                    "txnDate": normalize_date(row.get("Ngày tháng")),
                    "categoryName": normalize_text(row.get("Loại thu")) or normalize_text(row.get("Loại chi")),
                    "detail": normalize_text(row.get("Chi tiết các loại")),
                    "description": normalize_text(row.get("Diễn giải")),
                    "amount": normalize_number(row.get("Số tiền")) or 0,
                    "handledByName": normalize_text(row.get("Người thu/chi")),
                    "notes": normalize_text(row.get("Ghi chú")),
                    "billingPeriodRef": month_key(row.get("Tháng thu")) or month_key(row.get("Tháng chi")),
                }
            )

    summary["linkStats"] = {
        "studentsLinkedToLeads": sum(1 for item in students if item["leadRef"]),
        "studentGuardianLinks": len(student_guardian_links),
        "chargesResolvable": len(charges),
        "chargesUnresolved": unresolved_charge_links,
        "bookIssuesResolvable": len([item for item in book_issues if item["studentRef"]]),
        "bookIssuesUnresolved": unresolved_book_issue_links,
    }

    if unresolved_charge_links:
        summary["warnings"].append(
            f"Còn {unresolved_charge_links} dòng học phí chưa đủ khóa student/class/period để import charge chuẩn."
        )
    if unresolved_book_issue_links:
        summary["warnings"].append(
            f"Còn {unresolved_book_issue_links} dòng xuất sách chưa resolve chắc chắn sang học viên."
        )

    return {
        "lookups": {
            "courses": courses,
            "weekdays": raw_weekdays,
            "timeSlots": raw_time_slots,
        },
        "employees": employees,
        "crm": {
            "guardians": guardians,
            "leads": leads,
        },
        "academics": {
            "classes": classes,
            "students": students,
            "studentGuardianLinks": student_guardian_links,
            "enrollments": enrollments,
        },
        "finance": {
            "billingPeriods": billing_periods,
            "charges": charges,
            "payments": payments,
            "cashCategories": transaction_categories,
            "cashTransactions": cash_transactions,
        },
        "inventory": {
            "books": books,
            "stockReceipts": stock_receipts,
            "bookIssues": book_issues,
        },
        "summary": summary,
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def build_manifest(raw_tables: dict[str, list[dict[str, Any]]], canonical: dict[str, Any]) -> dict[str, Any]:
    raw_counts = {table: len(rows) for table, rows in raw_tables.items()}
    canonical_counts = {}
    for section_name, section_payload in canonical.items():
        if isinstance(section_payload, dict):
            for key, value in section_payload.items():
                if isinstance(value, list):
                    canonical_counts[f"{section_name}.{key}"] = len(value)
    return {
        "generatedAt": datetime.now().isoformat(),
        "rawTableCounts": raw_counts,
        "canonicalCounts": canonical_counts,
        "warnings": canonical["summary"]["warnings"],
        "linkStats": canonical["summary"]["linkStats"],
    }


def build_quality_markdown(diagnostics: dict[str, Any]) -> str:
    lines = [
        "# Workbook Data Quality Report",
        "",
        "## Tổng quan",
        "",
        f"- Tổng số bảng kiểm tra: {diagnostics['overall'].get('tables', 0)}",
        f"- Bảng sẵn sàng import: {diagnostics['overall'].get('readyTables', 0)}",
        f"- Bảng có khóa một phần: {diagnostics['overall'].get('partialTables', 0)}",
        f"- Bảng chủ yếu placeholder/công thức: {diagnostics['overall'].get('placeholderTables', 0)}",
        "",
        "## Chi tiết từng bảng",
        "",
    ]

    for table_name, report in diagnostics["tables"].items():
        lines.extend(
            [
                f"### {table_name}",
                f"- `qualityStatus`: {report['qualityStatus']}",
                f"- `rowCount`: {report['rowCount']}",
                f"- `businessKeyFields`: {', '.join(report['businessKeyFields']) if report['businessKeyFields'] else '(không cấu hình)'}",
                f"- `rowsWithAnyKey`: {report['rowsWithAnyKey']}",
                f"- `rowsWithAllKeys`: {report['rowsWithAllKeys']}",
                f"- `placeholderOnlyRows`: {report['placeholderOnlyRows']}",
                f"- `meaningfulRows`: {report['meaningfulRows']}",
            ]
        )
        if report["missingKeyCounts"]:
            lines.append("- `missingKeyCounts`:")
            for field, count in report["missingKeyCounts"].items():
                lines.append(f"  - {field}: {count}")
        if report["sampleRowsWithKeys"]:
            lines.append("- `sampleRowsWithKeys`:")
            for sample in report["sampleRowsWithKeys"]:
                lines.append(f"  - row {sample['sourceRow']}: {sample['keys']}")
        lines.append("")

    return "\n".join(lines)


def build_import_readiness(diagnostics: dict[str, Any], canonical: dict[str, Any]) -> dict[str, Any]:
    ready_tables = []
    partial_tables = []
    blocked_tables = []

    for table_name, report in diagnostics["tables"].items():
        item = {
            "table": table_name,
            "rowCount": report["rowCount"],
            "rowsWithAnyKey": report["rowsWithAnyKey"],
            "rowsWithAllKeys": report["rowsWithAllKeys"],
            "missingKeyCounts": report["missingKeyCounts"],
        }
        if report["qualityStatus"] == "READY_FOR_IMPORT":
            ready_tables.append(item)
        elif report["qualityStatus"] == "PARTIAL_KEYS":
            partial_tables.append(item)
        else:
            blocked_tables.append(item)

    ready_entities = {
        "courses": len(canonical["lookups"]["courses"]),
        "weekdays": len(canonical["lookups"]["weekdays"]),
        "timeSlots": len(canonical["lookups"]["timeSlots"]),
        "cashCategories": len(canonical["finance"]["cashCategories"]),
        "books": len(canonical["inventory"]["books"]),
        "bookIssues": len(canonical["inventory"]["bookIssues"]),
        "cashTransactions": len(canonical["finance"]["cashTransactions"]),
    }

    return {
        "readyTables": ready_tables,
        "partialTables": partial_tables,
        "blockedTables": blocked_tables,
        "readyEntities": ready_entities,
        "coreBlockers": [
            "DSTest/DSHV/DSLop/NhanSu/TheoDoiHP thiếu khóa nghiệp vụ ở hầu hết dòng.",
            "Thu-Chi.T_Thu và Thu-Chi.T_Chi hiện là vùng pivot/tổng hợp, không phải ledger raw.",
            "XuatNhapSach.T_SachXuat và T_SachNhap thiếu TenSach/NgayThang đầu vào để link chuẩn."
        ],
    }


def select_context_fields(table_name: str, rows: list[dict[str, Any]]) -> list[str]:
    key_fields = set(BUSINESS_KEY_FIELDS.get(table_name, []))
    counts = Counter()
    for row in rows:
        for field_name, value in row.items():
            if field_name == "__sourceRow" or field_name in key_fields:
                continue
            if not is_placeholder_value(value):
                counts[field_name] += 1
    return [field for field, _ in counts.most_common(6)]


def generate_remediation_templates(
    raw_tables: dict[str, list[dict[str, Any]]],
    diagnostics: dict[str, Any],
    remediation_dir: Path,
    refresh: bool,
) -> dict[str, Any]:
    remediation_dir.mkdir(parents=True, exist_ok=True)
    summary = {
        "filesWritten": 0,
        "tablesPrepared": 0,
    }

    index_rows = []
    excluded = {"MucLuc.Table2", "MucLuc.Table3", "Thu-Chi.T_PhanLoai", "XuatNhapSach.T_SachTon"}

    for table_name, report in diagnostics["tables"].items():
        if report["qualityStatus"] not in {"PLACEHOLDER_ONLY", "PARTIAL_KEYS"}:
            continue
        if table_name in excluded:
            continue

        rows = raw_tables.get(table_name, [])
        if not rows:
            continue

        business_fields = BUSINESS_KEY_FIELDS.get(table_name, [])
        context_fields = select_context_fields(table_name, rows)
        output_fields = ["sourceRow", "applyOverride", *business_fields, *context_fields, "notes"]
        output_path = remediation_dir / table_to_filename(table_name)

        if output_path.exists() and not refresh:
            index_rows.append({"table": table_name, "file": output_path.name, "status": "kept"})
            continue

        with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=output_fields)
            writer.writeheader()
            for row in rows:
                writer.writerow(
                    {
                        "sourceRow": row.get("__sourceRow"),
                        "applyOverride": "",
                        **{
                            field: "" if is_placeholder_value(row.get(field)) else row.get(field)
                            for field in business_fields + context_fields
                        },
                        "notes": "",
                    }
                )

        summary["filesWritten"] += 1
        summary["tablesPrepared"] += 1
        index_rows.append({"table": table_name, "file": output_path.name, "status": "written"})

    write_json(
        remediation_dir / "_index.json",
        {"generatedAt": datetime.now().isoformat(), "tables": index_rows},
    )
    return summary


def build_import_readiness_markdown(readiness: dict[str, Any]) -> str:
    lines = [
        "# Workbook Import Readiness",
        "",
        "## Import được ngay",
        "",
    ]

    for item in readiness["readyTables"]:
        lines.append(
            f"- `{item['table']}`: {item['rowsWithAllKeys']}/{item['rowCount']} dòng có đủ khóa"
        )

    lines.extend(
        [
            "",
            "## Cần bổ sung một phần",
            "",
        ]
    )
    for item in readiness["partialTables"]:
        lines.append(
            f"- `{item['table']}`: {item['rowsWithAnyKey']}/{item['rowCount']} dòng có khóa một phần"
        )

    lines.extend(
        [
            "",
            "## Đang bị chặn",
            "",
        ]
    )
    for item in readiness["blockedTables"]:
        lines.append(
            f"- `{item['table']}`: thiếu khóa trên {item['rowCount']} dòng"
        )

    lines.extend(
        [
            "",
            "## Blocker lõi",
            "",
        ]
    )
    for item in readiness["coreBlockers"]:
        lines.append(f"- {item}")

    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract workbook 2026 into ERP canonical JSON.")
    parser.add_argument("--source", default="docs/File Quan ly tong 2026.xlsx")
    parser.add_argument("--dictionary", default="docs/Data_Dictionary_Excel.csv")
    parser.add_argument("--output", default="docs/generated/workbook_2026")
    parser.add_argument("--remediation-dir", default=None)
    parser.add_argument("--refresh-remediation", action="store_true")
    args = parser.parse_args()

    source_path = Path(args.source)
    dictionary_path = Path(args.dictionary)
    output_path = Path(args.output)
    remediation_dir = Path(args.remediation_dir) if args.remediation_dir else output_path / "remediation"

    specs = read_dictionary(dictionary_path)
    field_type_lookup = build_field_type_lookup(specs)
    raw_tables_extracted = extract_tables(source_path, specs)
    csv_overrides, override_load_summary = load_csv_overrides(remediation_dir, field_type_lookup)
    raw_tables, override_apply_summary = apply_overrides(raw_tables_extracted, csv_overrides)
    canonical = canonicalize(raw_tables)
    diagnostics = build_diagnostics(raw_tables)
    manifest = build_manifest(raw_tables, canonical)
    readiness = build_import_readiness(diagnostics, canonical)
    remediation_summary = generate_remediation_templates(
        raw_tables,
        diagnostics,
        remediation_dir,
        args.refresh_remediation,
    )

    write_json(output_path / "raw_tables_extracted.json", raw_tables_extracted)
    write_json(output_path / "raw_tables.json", raw_tables)
    write_json(output_path / "canonical.json", canonical)
    write_json(output_path / "diagnostics.json", diagnostics)
    write_json(output_path / "manifest.json", manifest)
    write_json(output_path / "import_readiness.json", readiness)
    write_json(
        output_path / "override_summary.json",
        {
            "loaded": override_load_summary,
            "applied": override_apply_summary,
            "remediationTemplates": remediation_summary,
        },
    )
    (output_path / "data_quality_report.md").write_text(
        build_quality_markdown(diagnostics),
        encoding="utf-8",
    )
    (output_path / "import_readiness.md").write_text(
        build_import_readiness_markdown(readiness),
        encoding="utf-8",
    )

    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
